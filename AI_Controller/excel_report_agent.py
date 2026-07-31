import os
import sqlite3
import duckdb
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define paths
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DUCKDB_PATH = os.path.join(BASE_DIR, "Data", "DuckDB", "project_controlling.db")
SQLITE_PATH = os.path.join(BASE_DIR, "Data", "SQLite", "project_controlling.db")
REPORT_PATH = os.path.join(BASE_DIR, "Data", "vessel_construction_report.xlsx")

def create_excel_report():
    print("Loading data from DuckDB...")
    con_dd = duckdb.connect(DUCKDB_PATH, read_only=True)
    
    # 1. EVM Summary
    summary = con_dd.execute("""
        SELECT Total_BAC, Total_AC, Total_EV, Total_CV, Project_CPI, Total_EAC_Typical, Total_VAC, Overall_Progress_Pct 
        FROM v_project_evm_summary
    """).fetchone()
    
    # 2. WBS Metrics
    wbs_metrics = con_dd.execute("""
        SELECT WBS_Code, ElementName, BAC, AC, EV, CPI, PercentComplete, EAC_Typical 
        FROM v_wbs_evm_metrics
        ORDER BY WBS_Code
    """).fetchall()
    
    # 3. Labor Details
    labor_summary = con_dd.execute("""
        SELECT r.ResourceName, r.Role, r.HourlyRate, SUM(t.HoursWorked) as TotalHours, SUM(t.HoursWorked * r.HourlyRate) as TotalCost
        FROM timesheets t
        JOIN resources r ON t.ResourceID = r.ResourceID
        GROUP BY r.ResourceName, r.Role, r.HourlyRate
        ORDER BY TotalCost DESC
    """).fetchall()
    
    # 4. Materials Details
    materials_summary = con_dd.execute("""
        SELECT PurchaseDate, PurchaseID, Description, TotalActualCost
        FROM material_costs
        ORDER BY PurchaseDate
    """).fetchall()
    
    con_dd.close()
    
    print("Loading data from SQLite...")
    con_sq = sqlite3.connect(SQLITE_PATH)
    
    # 5. Overtime
    overtime_logs = con_sq.execute("""
        SELECT strftime('%Y-%W', t.WorkDate) as WorkWeek, r.ResourceName, r.Role, SUM(CAST(t.HoursWorked AS REAL)) as TotalHours
        FROM timesheets t
        JOIN resources r ON t.ResourceID = r.ResourceID
        GROUP BY WorkWeek, r.ResourceName, r.Role
        HAVING TotalHours > 45
        ORDER BY WorkWeek DESC, TotalHours DESC
    """).fetchall()
    
    # 6. RAID Logs
    raid_logs = con_sq.execute("""
        SELECT RiskID, Type, Description, Impact, Probability, MitigationStrategy, Owner, Status
        FROM raid_log
        ORDER BY RiskID DESC
    """).fetchall()
    
    con_sq.close()
    
    # Initialize OpenPyXL Workbook
    wb = openpyxl.Workbook()
    
    # Common Styles
    title_font = Font(name="Segoe UI", size=16, bold=True, color="2C3E50")
    subtitle_font = Font(name="Segoe UI", size=11, italic=True, color="7F8C8D")
    section_font = Font(name="Segoe UI", size=13, bold=True, color="1A252F")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Segoe UI", size=11, bold=True, color="2C3E50")
    regular_font = Font(name="Segoe UI", size=11, color="333333")
    muted_font = Font(name="Segoe UI", size=10, italic=True, color="7F8C8D")
    
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    accent_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    alert_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    ok_fill = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid")
    
    thin_bottom = Border(bottom=Side(style='thin', color='BDC3C7'))
    all_thin = Border(
        left=Side(style='thin', color='E5E7E9'),
        right=Side(style='thin', color='E5E7E9'),
        top=Side(style='thin', color='E5E7E9'),
        bottom=Side(style='thin', color='E5E7E9')
    )
    double_bottom = Border(bottom=Side(style='double', color='2C3E50'), top=Side(style='thin', color='BDC3C7'))
    
    # ====================================================
    # TAB 1: Control Tower Dashboard
    # ====================================================
    ws_dash = wb.active
    ws_dash.title = "Control Tower Dashboard"
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_dash["A1"] = "⚓ PROJECT CONTROL TOWER DASHBOARD"
    ws_dash["A1"].font = title_font
    ws_dash["A2"] = "Relational EVM & Analytical Controller Dashboard | PRINCE2 Compliance"
    ws_dash["A2"].font = subtitle_font
    ws_dash.row_dimensions[1].height = 28
    
    # KPI metrics (BAC, AC, EV, CPI, Progress, CV, VAC, TCPI)
    kpis = [
        ("BUDGET (BAC)", summary[0], "$#,##0.00"),
        ("ACTUAL COST (AC)", summary[1], "$#,##0.00"),
        ("EARNED VALUE (EV)", summary[2], "$#,##0.00"),
        ("CPI", summary[4], "0.00"),
        ("PROGRESS", summary[7] / 100, "0.0%")
    ]
    
    for i, (label, val, fmt) in enumerate(kpis):
        col_idx = i * 2 + 1
        ws_dash.cell(row=4, column=col_idx, value=label).font = muted_font
        val_cell = ws_dash.cell(row=5, column=col_idx, value=val)
        val_cell.font = Font(name="Segoe UI", size=18, bold=True, color="2C3E50")
        val_cell.number_format = fmt
        if label == "CPI" and val < 0.95:
            val_cell.fill = alert_fill
            
    # WBS Table Headers
    ws_dash.cell(row=8, column=1, value="WBS Element Performance").font = section_font
    headers = ["WBS Code", "ElementName", "BAC (USD)", "AC (USD)", "EV (USD)", "CPI", "Progress", "EAC Typical (USD)", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_dash.cell(row=9, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in [1, 9] else ("left" if col_idx == 2 else "right"))
    ws_dash.row_dimensions[9].height = 22
    
    start_row = 10
    for i, row in enumerate(wbs_metrics):
        r = start_row + i
        wbs_code, name, bac, ac, ev, cpi, pct, eac = row
        
        ws_dash.cell(row=r, column=1, value=wbs_code).alignment = Alignment(horizontal="center")
        ws_dash.cell(row=r, column=2, value=name).alignment = Alignment(horizontal="left")
        ws_dash.cell(row=r, column=3, value=bac).number_format = "$#,##0.00"
        ws_dash.cell(row=r, column=4, value=ac).number_format = "$#,##0.00"
        ws_dash.cell(row=r, column=5, value=ev).number_format = "$#,##0.00"
        
        cpi_cell = ws_dash.cell(row=r, column=6, value=f"=E{r}/D{r}")
        cpi_cell.number_format = "0.00"
        cpi_cell.alignment = Alignment(horizontal="right")
        
        ws_dash.cell(row=r, column=7, value=pct).number_format = "0.0%"
        
        eac_cell = ws_dash.cell(row=r, column=8, value=f"=C{r}/F{r}")
        eac_cell.number_format = "$#,##0.00"
        
        status_cell = ws_dash.cell(row=r, column=9, value=f'=IF(F{r}<0.95, "🔴 Overrun", "🟢 On Track")')
        status_cell.alignment = Alignment(horizontal="center")
        
        for c in range(1, 10):
            cell = ws_dash.cell(row=r, column=c)
            cell.font = regular_font
            cell.border = thin_bottom
            if cpi < 0.95 and c in [6, 9]:
                cell.fill = alert_fill
                
    tot_row = start_row + len(wbs_metrics)
    ws_dash.cell(row=tot_row, column=1, value="TOTAL").alignment = Alignment(horizontal="center")
    ws_dash.cell(row=tot_row, column=2, value="Project Vessel Summary")
    ws_dash.cell(row=tot_row, column=3, value=f"=SUM(C10:C{tot_row-1})").number_format = "$#,##0.00"
    ws_dash.cell(row=tot_row, column=4, value=f"=SUM(D10:D{tot_row-1})").number_format = "$#,##0.00"
    ws_dash.cell(row=tot_row, column=5, value=f"=SUM(E10:E{tot_row-1})").number_format = "$#,##0.00"
    ws_dash.cell(row=tot_row, column=6, value=f"=E{tot_row}/D{tot_row}").number_format = "0.00"
    ws_dash.cell(row=tot_row, column=7, value=f"=E{tot_row}/C{tot_row}").number_format = "0.0%"
    ws_dash.cell(row=tot_row, column=8, value=f"=SUM(H10:H{tot_row-1})").number_format = "$#,##0.00"
    ws_dash.cell(row=tot_row, column=9, value=f'=IF(F{tot_row}<0.95, "🔴 Overrun", "🟢 On Track")').alignment = Alignment(horizontal="center")
    
    for c in range(1, 10):
        cell = ws_dash.cell(row=tot_row, column=c)
        cell.font = bold_font
        cell.border = double_bottom
        cell.fill = accent_fill
        
    # ====================================================
    # TAB 2: Schedule & Gantt
    # ====================================================
    ws_gantt = wb.create_sheet(title="Schedule & Gantt")
    ws_gantt.views.sheetView[0].showGridLines = True
    
    ws_gantt["A1"] = "📅 PROJECT Gantt CHART & SCHEDULE"
    ws_gantt["A1"].font = title_font
    ws_gantt["A2"] = "Chris Croft Method: Durations, Predecessors, Float & Critical Path Identification"
    ws_gantt["A2"].font = subtitle_font
    
    gantt_headers = [
        "WBS", "Task Description", "Start Date", "End Date", 
        "Duration (Days)", "Predecessors", "Total Float", "Critical?", 
        "Jan 2026", "Feb 2026", "Mar 2026", "Apr 2026", "May 2026", "Jun 2026"
    ]
    
    for col_idx, header in enumerate(gantt_headers, start=1):
        cell = ws_gantt.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws_gantt.row_dimensions[5].height = 22
    
    tasks = [
        ("1.0", "Project Management & Eng.", "2026-01-01", "2026-06-30", 181, "None", 0, "Yes", [1, 1, 1, 1, 1, 1]),
        ("2.0", "Hull Fabrication & Assembly", "2026-02-01", "2026-04-15", 74, "1.0", 0, "Yes", [0, 1, 1, 1, 0, 0]),
        ("3.0", "Outfitting & Integration", "2026-04-01", "2026-05-31", 61, "2.0", 0, "Yes", [0, 0, 0, 1, 1, 0]),
        ("4.0", "Sea Trials & Handover", "2026-06-01", "2026-06-30", 30, "3.0", 0, "Yes", [0, 0, 0, 0, 0, 1]),
    ]
    
    critical_fill = PatternFill(start_color="F2D7D5", end_color="F2D7D5", fill_type="solid") # soft red
    pm_fill = PatternFill(start_color="EAEDED", end_color="EAEDED", fill_type="solid") # soft gray
    
    for i, task in enumerate(tasks):
        r = 6 + i
        wbs_code, name, start, end, duration, pred, float_val, is_crit, months = task
        
        ws_gantt.cell(row=r, column=1, value=wbs_code).alignment = Alignment(horizontal="center")
        ws_gantt.cell(row=r, column=2, value=name).alignment = Alignment(horizontal="left")
        ws_gantt.cell(row=r, column=3, value=start).alignment = Alignment(horizontal="center")
        ws_gantt.cell(row=r, column=4, value=end).alignment = Alignment(horizontal="center")
        ws_gantt.cell(row=r, column=5, value=duration).alignment = Alignment(horizontal="right")
        ws_gantt.cell(row=r, column=6, value=pred).alignment = Alignment(horizontal="center")
        ws_gantt.cell(row=r, column=7, value=float_val).alignment = Alignment(horizontal="right")
        
        crit_cell = ws_gantt.cell(row=r, column=8, value=is_crit)
        crit_cell.alignment = Alignment(horizontal="center")
        if is_crit == "Yes":
            crit_cell.font = Font(name="Segoe UI", bold=True, color="922B21")
            
        for m_idx, active in enumerate(months):
            col = 9 + m_idx
            bar_cell = ws_gantt.cell(row=r, column=col, value="" if not active else "■")
            bar_cell.alignment = Alignment(horizontal="center")
            if active:
                if wbs_code == "1.0":
                    bar_cell.fill = pm_fill
                    bar_cell.font = Font(color="7F8C8D", bold=True)
                else:
                    bar_cell.fill = critical_fill
                    bar_cell.font = Font(color="922B21", bold=True)
                    
        for c in range(1, 15):
            cell = ws_gantt.cell(row=r, column=c)
            cell.font = regular_font
            if c <= 8:
                cell.border = thin_bottom
            else:
                cell.border = all_thin

    # ====================================================
    # TAB 3: CFO & Cost Share Audit
    # ====================================================
    ws_cfo = wb.create_sheet(title="CFO & Cost Share Audit")
    ws_cfo.views.sheetView[0].showGridLines = True
    
    ws_cfo["A1"] = "💼 CFO PROFITABILITY & COST SHARE AUDIT"
    ws_cfo["A1"].font = title_font
    ws_cfo["A2"] = "Resource and material allocations, burn rates, and project margins"
    ws_cfo["A2"].font = subtitle_font
    
    # Labor vs Material Split table
    ws_cfo.cell(row=5, column=1, value="Cost-Share Breakdown").font = section_font
    cfo_headers = ["Category", "Actual Cost (USD)", "Share %"]
    for col_idx, h in enumerate(cfo_headers, start=1):
        cell = ws_cfo.cell(row=6, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "right")
        
    tot_labor = sum(row[4] for row in labor_summary)
    tot_material = sum(row[3] for row in materials_summary)
    
    # Write Labor
    ws_cfo.cell(row=7, column=1, value="Labor Costs").font = regular_font
    ws_cfo.cell(row=7, column=2, value=tot_labor).number_format = "$#,##0.00"
    ws_cfo.cell(row=7, column=3, value=f"=B7/B9").number_format = "0.0%"
    
    # Write Material
    ws_cfo.cell(row=8, column=1, value="Material & Procurement Costs").font = regular_font
    ws_cfo.cell(row=8, column=2, value=tot_material).number_format = "$#,##0.00"
    ws_cfo.cell(row=8, column=3, value=f"=B8/B9").number_format = "0.0%"
    
    # Total Cost Share
    ws_cfo.cell(row=9, column=1, value="TOTAL ACTUAL COST").font = bold_font
    ws_cfo.cell(row=9, column=2, value=f"=SUM(B7:B8)").number_format = "$#,##0.00"
    ws_cfo.cell(row=9, column=3, value=1.0).number_format = "0.0%"
    
    for c in range(1, 4):
        ws_cfo.cell(row=7, column=c).border = thin_bottom
        ws_cfo.cell(row=8, column=c).border = thin_bottom
        tot_cell = ws_cfo.cell(row=9, column=c)
        tot_cell.font = bold_font
        tot_cell.fill = accent_fill
        tot_cell.border = double_bottom
        
    # Top Resource Burn Rates
    ws_cfo.cell(row=12, column=1, value="Resource Burn Rates").font = section_font
    res_headers = ["Resource Name", "Role", "Hourly Rate (USD)", "Hours Logged", "Total Cost Logged (USD)"]
    for col_idx, h in enumerate(res_headers, start=1):
        cell = ws_cfo.cell(row=13, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left" if col_idx <= 2 else "right")
        
    for idx, row in enumerate(labor_summary):
        r = 14 + idx
        name, role, rate, hours, cost = row
        ws_cfo.cell(row=r, column=1, value=name).font = regular_font
        ws_cfo.cell(row=r, column=2, value=role).font = regular_font
        ws_cfo.cell(row=r, column=3, value=float(rate) * 0.10).number_format = "$#,##0.00"
        ws_cfo.cell(row=r, column=4, value=float(hours)).number_format = "#,##0.0"
        ws_cfo.cell(row=r, column=5, value=float(cost)).number_format = "$#,##0.00"
        for c in range(1, 6):
            ws_cfo.cell(row=r, column=c).border = thin_bottom
            
    # ====================================================
    # TAB 4: Contract & RAID Risk Audit
    # ====================================================
    ws_risk = wb.create_sheet(title="Contract & RAID Risk Audit")
    ws_risk.views.sheetView[0].showGridLines = True
    
    ws_risk["A1"] = "🚨 CONTRACT, RISK & ANOMALY AUDIT"
    ws_risk["A1"].font = title_font
    ws_risk["A2"] = "Resource weekly overtime limits, high-value invoices, and active RAID registry"
    ws_risk["A2"].font = subtitle_font
    
    # Overtime Audit Table
    ws_risk.cell(row=5, column=1, value="Overtime Audits (>45 Hours/Week)").font = section_font
    ot_headers = ["Work Week", "Resource Name", "Role", "Hours Logged"]
    for col_idx, h in enumerate(ot_headers, start=1):
        cell = ws_risk.cell(row=6, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx == 1 else ("left" if col_idx in [2, 3] else "right"))
        
    for idx, row in enumerate(overtime_logs):
        r = 7 + idx
        week, name, role, hours = row
        ws_risk.cell(row=r, column=1, value=week).alignment = Alignment(horizontal="center")
        ws_risk.cell(row=r, column=2, value=name).alignment = Alignment(horizontal="left")
        ws_risk.cell(row=r, column=3, value=role).alignment = Alignment(horizontal="left")
        h_cell = ws_risk.cell(row=r, column=4, value=float(hours))
        h_cell.number_format = "#,##0.0"
        h_cell.alignment = Alignment(horizontal="right")
        h_cell.fill = alert_fill
        for c in range(1, 5):
            ws_risk.cell(row=r, column=c).border = thin_bottom
            ws_risk.cell(row=r, column=c).font = regular_font
            
    # Large Procurement Transactions Table
    start_proc_row = 7 + len(overtime_logs) + 3
    ws_risk.cell(row=start_proc_row-1, column=1, value="Large Procurement Audits (>$5,000)").font = section_font
    proc_headers = ["Purchase Date", "Invoice Number", "Description", "Cost (USD)"]
    for col_idx, h in enumerate(proc_headers, start=1):
        cell = ws_risk.cell(row=start_proc_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in [1, 2] else ("left" if col_idx == 3 else "right"))
        
    large_invoices = [inv for inv in materials_summary if inv[3] > 5000]
    r = start_proc_row
    for idx, inv in enumerate(large_invoices):
        r = start_proc_row + 1 + idx
        date_str, inv_id, desc, cost = inv
        ws_risk.cell(row=r, column=1, value=date_str).alignment = Alignment(horizontal="center")
        ws_risk.cell(row=r, column=2, value=inv_id).alignment = Alignment(horizontal="center")
        ws_risk.cell(row=r, column=3, value=desc).alignment = Alignment(horizontal="left")
        cost_cell = ws_risk.cell(row=r, column=4, value=float(cost))
        cost_cell.number_format = "$#,#0.00"
        cost_cell.alignment = Alignment(horizontal="right")
        cost_cell.fill = alert_fill
        for c in range(1, 5):
            ws_risk.cell(row=r, column=c).border = thin_bottom
            ws_risk.cell(row=r, column=c).font = regular_font

    # Active RAID Log Table
    start_raid_row = r + 3
    ws_risk.cell(row=start_raid_row-1, column=1, value="Active RAID Log Register").font = section_font
    raid_headers = ["RAID ID", "Category", "Description", "Impact", "Probability", "Mitigation Strategy", "Owner", "Status"]
    for col_idx, h in enumerate(raid_headers, start=1):
        cell = ws_risk.cell(row=start_raid_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in [1, 4, 5, 8] else "left")
        
    active_raids = [item for item in raid_logs if item[7] == 'Active']
    for idx, item in enumerate(active_raids):
        r = start_raid_row + 1 + idx
        rid, cat, desc, impact, prob, mit, owner, status = item
        ws_risk.cell(row=r, column=1, value=rid).alignment = Alignment(horizontal="center")
        ws_risk.cell(row=r, column=2, value=cat).alignment = Alignment(horizontal="left")
        ws_risk.cell(row=r, column=3, value=desc).alignment = Alignment(horizontal="left")
        ws_risk.cell(row=r, column=4, value=impact).alignment = Alignment(horizontal="center")
        ws_risk.cell(row=r, column=5, value=prob).alignment = Alignment(horizontal="center")
        ws_risk.cell(row=r, column=6, value=mit).alignment = Alignment(horizontal="left")
        ws_risk.cell(row=r, column=7, value=owner).alignment = Alignment(horizontal="left")
        ws_risk.cell(row=r, column=8, value=status).alignment = Alignment(horizontal="center")
        for c in range(1, 9):
            ws_risk.cell(row=r, column=c).border = thin_bottom
            ws_risk.cell(row=r, column=c).font = regular_font
            if impact == 'High':
                ws_risk.cell(row=r, column=c).fill = alert_fill

    # ====================================================
    # TAB 5: Interactive Scenario Simulator
    # ====================================================
    ws_sim = wb.create_sheet(title="Interactive Simulator")
    ws_sim.views.sheetView[0].showGridLines = True
    
    ws_sim["A1"] = "🎛️ WHAT-IF CORRECTIVE ACTION & CRASH SIMULATOR"
    ws_sim["A1"].font = title_font
    ws_sim["A2"] = "Edit the yellow input cells to simulate real-time corrective actions, rate savings, and schedule crashing"
    ws_sim["A2"].font = subtitle_font
    
    # Simulator Input Section
    ws_sim.cell(row=5, column=1, value="Simulator Inputs (Editable)").font = section_font
    
    inputs = [
        ("Simulated Labor Rate Savings (%)", 0.05, "0.0%"),
        ("Simulated Material Price Savings (%)", 0.10, "0.0%"),
        ("Contract Delay Penalty (USD/Day)", 1000.0, "$#,##0.00"),
        ("Crash Schedule? (0=No, 5=Crash 5 Days, 10=Crash 10 Days)", 5, "0")
    ]
    
    input_fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid") # light yellow
    
    for idx, (label, val, fmt) in enumerate(inputs):
        r = 6 + idx
        ws_sim.cell(row=r, column=1, value=label).font = bold_font
        val_cell = ws_sim.cell(row=r, column=2, value=val)
        val_cell.font = bold_font
        val_cell.fill = input_fill
        val_cell.number_format = fmt
        val_cell.alignment = Alignment(horizontal="right")
        val_cell.border = all_thin
        
    # Simulator Outputs Section
    ws_sim.cell(row=12, column=1, value="Simulation Results (Calculated)").font = section_font
    
    # Original Values
    ws_sim.cell(row=14, column=1, value="Current Baseline BAC (USD)").font = regular_font
    ws_sim.cell(row=14, column=2, value=summary[0]).number_format = "$#,##0.00"
    
    ws_sim.cell(row=15, column=1, value="Current Actual Cost AC (USD)").font = regular_font
    ws_sim.cell(row=15, column=2, value=summary[1]).number_format = "$#,##0.00"
    
    ws_sim.cell(row=16, column=1, value="Current Earned Value EV (USD)").font = regular_font
    ws_sim.cell(row=16, column=2, value=summary[2]).number_format = "$#,##0.00"
    
    # Formulas for simulated cost share
    ws_sim.cell(row=18, column=1, value="Simulated Actual Cost (AC) USD").font = bold_font
    # AC = Current_AC * (1 - Labor_Saving * 0.718 - Material_Saving * 0.282) + Crashing_Cost
    ac_form = "=B15 * (1 - B6 * 0.718 - B7 * 0.282) + 20000 * (B9 / 100)"
    ws_sim.cell(row=18, column=2, value=ac_form).number_format = "$#,##0.00"
    ws_sim.cell(row=18, column=2).font = bold_font
    ws_sim.cell(row=18, column=2).fill = accent_fill
    
    # Simulated CPI
    ws_sim.cell(row=19, column=1, value="Simulated CPI").font = bold_font
    ws_sim.cell(row=19, column=2, value="=B16/B18").number_format = "0.00"
    ws_sim.cell(row=19, column=2).font = bold_font
    
    # Simulated EAC
    ws_sim.cell(row=20, column=1, value="Simulated Estimate At Completion (EAC) USD").font = bold_font
    ws_sim.cell(row=20, column=2, value="=B14/B19").number_format = "$#,##0.00"
    ws_sim.cell(row=20, column=2).font = bold_font
    ws_sim.cell(row=20, column=2).fill = accent_fill
    
    # Simulated Delay & Liquidated Damages
    ws_sim.cell(row=22, column=1, value="Simulated Schedule Delay (Days)").font = regular_font
    ws_sim.cell(row=22, column=2, value="=MAX(0, 10 - B9)").number_format = "0"
    
    ws_sim.cell(row=23, column=1, value="Simulated Liquidated Damages Penalty (USD)").font = regular_font
    ws_sim.cell(row=23, column=2, value="=B22 * B8").number_format = "$#,##0.00"
    
    ws_sim.cell(row=24, column=1, value="Direct Schedule Crashing Cost (USD)").font = regular_font
    ws_sim.cell(row=24, column=2, value="=20000 * (B9 / 100)").number_format = "$#,##0.00"
    
    ws_sim.cell(row=25, column=1, value="Net Crashing Trade-off Benefit (USD)").font = bold_font
    ws_sim.cell(row=25, column=2, value="=(10 * B8 - B23) - B24").number_format = "$#,##0.00"
    ws_sim.cell(row=25, column=2).font = bold_font
    ws_sim.cell(row=25, column=2).fill = ok_fill
    
    # Borders for Simulator
    for r in range(14, 26):
        ws_sim.cell(row=r, column=1).border = thin_bottom
        ws_sim.cell(row=r, column=2).border = thin_bottom
        
    # Auto-adjust column widths
    for ws in [ws_dash, ws_gantt, ws_cfo, ws_risk, ws_sim]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if val_str.startswith('='): # Don't size based on formulas
                    val_str = "$184,881.00"
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
    # Save Workbook
    wb.save(REPORT_PATH)
    print(f"Excel report successfully generated and saved to: {REPORT_PATH}")

if __name__ == "__main__":
    create_excel_report()
