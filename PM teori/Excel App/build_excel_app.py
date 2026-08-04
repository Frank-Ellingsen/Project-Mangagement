import os
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define directories
BASE_DIR = r"c:\Users\frank\Desktop\Project Mng\PM teori\Excel App"
CSV_DIR = os.path.join(BASE_DIR, "CSV")
OUTPUT_PATH = os.path.join(BASE_DIR, "Project_Controlling_App.xlsx")

def build_workbook():
    wb = openpyxl.Workbook()
    
    # Define Tufte & clean styles
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="2C3E50")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="7F8C8D")
    section_font = Font(name=font_family, size=12, bold=True, color="2C3E50")
    header_font = Font(name=font_family, size=10, bold=True, color="333333")
    data_font = Font(name=font_family, size=10, color="000000")
    bold_data_font = Font(name=font_family, size=10, bold=True, color="000000")
    instruction_font = Font(name=font_family, size=9, italic=True, color="555555")
    
    # Table Header and Total Fills
    header_fill = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
    total_fill = PatternFill(start_color="FBFCFC", end_color="FBFCFC", fill_type="solid")
    
    # Alert Fills (Muted/Soft red and amber to adhere to Tufte principles)
    alert_fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid") # soft red
    warn_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")  # soft yellow/amber
    
    # Borders
    thin_gray = Side(style='thin', color='E5E7E9')
    medium_dark = Side(style='thin', color='BDC3C7')
    double_dark = Side(style='double', color='2C3E50')
    
    border_header = Border(bottom=medium_dark, top=thin_gray)
    border_data = Border(bottom=thin_gray)
    border_total = Border(top=medium_dark, bottom=double_dark)
    
    # ----------------------------------------------------
    # Load and Staging CSVs
    # ----------------------------------------------------
    csv_files = [
        ("Dim_Project.csv", "Dim_Project"),
        ("Dim_WBS.csv", "Dim_WBS"),
        ("Dim_Resource.csv", "Dim_Resource"),
        ("Dim_Calendar.csv", "Dim_Calendar"),
        ("Fact_Baseline_PV.csv", "Fact_Baseline_PV"),
        ("Fact_Actual_Costs.csv", "Fact_Actual_Costs"),
        ("Fact_Physical_Progress.csv", "Fact_Physical_Progress"),
    ]
    
    from datetime import datetime
    
    for filename, sheetname in csv_files:
        filepath = os.path.join(CSV_DIR, filename)
        ws = wb.create_sheet(title=sheetname)
        ws.views.sheetView[0].showGridLines = True
        
        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, start=1):
                for col_idx, val in enumerate(row, start=1):
                    # Convert types where appropriate
                    try:
                        if len(val) == 10 and val[4] == '-' and val[7] == '-':
                            typed_val = datetime.strptime(val, "%Y-%m-%d").date()
                        elif "." in val and not val.replace(".", "", 1).isdigit():
                            typed_val = val
                        elif val.isdigit():
                            typed_val = int(val)
                        else:
                            typed_val = float(val)
                    except ValueError:
                        typed_val = val
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=typed_val)
                    cell.font = data_font
                    
                    # If it's a date cell, set formatting
                    if isinstance(typed_val, (datetime, datetime.date.__class__, type(datetime.now().date()))):
                        cell.number_format = "yyyy-mm-dd"
                    
                    # Format headers
                    if row_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = border_header
                        
        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # ----------------------------------------------------
    # Build EVM Dashboard
    # ----------------------------------------------------
    ws_dash = wb.active
    ws_dash.title = "EVM Dashboard"
    ws_dash.views.sheetView[0].showGridLines = False # Erase gridlines for Tufte styling
    
    # Title Block
    ws_dash["A1"] = "OFFICE FIT-OUT - PROJECT CONTROLLING DASHBOARD"
    ws_dash["A1"].font = title_font
    ws_dash["A2"] = "EVM Performance Status | Tufte-Compliant Design | Data-Ink Maximized"
    ws_dash["A2"].font = subtitle_font
    ws_dash.row_dimensions[1].height = 24
    ws_dash.row_dimensions[2].height = 18
    
    # Metadata Block
    ws_dash["A4"] = "Status Date:"
    ws_dash["A4"].font = bold_data_font
    ws_dash["B4"] = datetime.strptime("2026-06-12", "%Y-%m-%d").date()
    ws_dash["B4"].font = data_font
    ws_dash["B4"].number_format = "yyyy-mm-dd"
    ws_dash["B4"].alignment = Alignment(horizontal="left")
    
    ws_dash["D4"] = "Status Week:"
    ws_dash["D4"].font = bold_data_font
    ws_dash["E4"] = 6
    ws_dash["E4"].font = data_font
    ws_dash["E4"].alignment = Alignment(horizontal="left")
    
    ws_dash["G4"] = "Sector:"
    ws_dash["G4"].font = bold_data_font
    ws_dash["H4"] = "Corporate Real Estate"
    ws_dash["H4"].font = data_font
    
    # ----------------------------------------------------
    # KPI Summary Cards Block (Rows 6-13)
    # ----------------------------------------------------
    ws_dash["A6"] = "Executive Metrics Summary"
    ws_dash["A6"].font = section_font
    
    kpis = [
        # Label, Formula, Format, ColStart, RowIdx
        ("Budget at Completion (BAC)", "=D30", "#,##0.00\" NOK\"", "A", 7),
        ("Planned Value (PV)", "=E30", "#,##0.00\" NOK\"", "D", 7),
        ("Actual Cost (AC)", "=F30", "#,##0.00\" NOK\"", "G", 7),
        ("Earned Value (EV)", "=H30", "#,##0.00\" NOK\"", "J", 7),
        
        ("Cost Variance (CV)", "=I30", "#,##0.00\" NOK\"", "A", 10),
        ("Schedule Variance (SV)", "=J30", "#,##0.00\" NOK\"", "D", 10),
        ("CPI (Cost Index)", "=K30", "0.00", "G", 10),
        ("SPI (Schedule Index)", "=L30", "0.00", "J", 10),
        
        ("Estimate at Completion (EAC)", "=M30", "#,##0.00\" NOK\"", "A", 13),
        ("Estimate to Complete (ETC)", "=N30", "#,##0.00\" NOK\"", "D", 13),
        ("Variance at Completion (VAC)", "=O30", "#,##0.00\" NOK\"", "G", 13),
        ("Project Progress %", "=G30", "0.0%", "J", 13)
    ]
    
    # Draw KPI cells
    for label, formula, num_format, col, row_idx in kpis:
            
        col_letter = col
        # Find next col to merge
        next_col = chr(ord(col_letter) + 2)
        
        # Label cell
        lbl_cell = ws_dash[f"{col_letter}{row_idx-1}"]
        lbl_cell.value = label.upper()
        lbl_cell.font = Font(name=font_family, size=9, bold=True, color="7F8C8D")
        
        # Value cell
        val_cell = ws_dash[f"{col_letter}{row_idx}"]
        val_cell.value = formula
        val_cell.font = Font(name=font_family, size=12, bold=True, color="2C3E50")
        val_cell.number_format = num_format
        val_cell.alignment = Alignment(horizontal="left", vertical="center")
        val_cell.border = Border(bottom=thin_gray)
        val_cell.fill = total_fill
        
        # Merge columns to make clean boxes
        ws_dash.merge_cells(f"{col_letter}{row_idx}:{next_col}{row_idx}")
        
    ws_dash.row_dimensions[7].height = 24
    ws_dash.row_dimensions[10].height = 24
    ws_dash.row_dimensions[13].height = 24
    
    # ----------------------------------------------------
    # WBS Performance Table (Row 15+)
    # ----------------------------------------------------
    ws_dash["A15"] = "WBS Element Performance Breakdown"
    ws_dash["A15"].font = section_font
    
    headers = [
        "WBS ID", "Activity Name", "Type", "Budget (BAC)", "Planned Value (PV)", 
        "Actual Cost (AC)", "Progress %", "Earned Value (EV)", "CV", "SV", 
        "CPI", "SPI", "EAC", "ETC", "VAC", "Status"
    ]
    
    for idx, h in enumerate(headers, start=1):
        cell = ws_dash.cell(row=17, column=idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_header
        cell.alignment = Alignment(horizontal="left" if idx in [1, 2, 3] else ("center" if idx == 16 else "right"))
        
    # Populate activities (rows 18 to 29)
    wbs_activities = [
        ("WBS-01", "Project Management", "Labor"),
        ("WBS-02", "Site Survey & Mobilization", "Labor"),
        ("WBS-03", "Design & Engineering", "Labor"),
        ("WBS-04", "Procurement - Materials", "Material"),
        ("WBS-05", "Demolition & Prep", "Labor"),
        ("WBS-06", "Electrical Rough-In", "Labor"),
        ("WBS-07", "HVAC Installation", "Subcontractor"),
        ("WBS-08", "Partitions & Drywall", "Subcontractor"),
        ("WBS-09", "Flooring", "Material"),
        ("WBS-10", "Fixtures & Furniture", "Material"),
        ("WBS-11", "Testing / Commissioning", "Labor"),
        ("WBS-12", "Handover & Closeout", "Labor")
    ]
    
    for i, (wbs_id, name, w_type) in enumerate(wbs_activities, start=18):
        # WBS ID, Name, Type
        ws_dash.cell(row=i, column=1, value=wbs_id).alignment = Alignment(horizontal="left")
        ws_dash.cell(row=i, column=2, value=name).alignment = Alignment(horizontal="left")
        ws_dash.cell(row=i, column=3, value=w_type).alignment = Alignment(horizontal="left")
        
        # Budget (BAC) - Pull from Dim_WBS
        ws_dash.cell(row=i, column=4, value=f"=VLOOKUP(A{i}, Dim_WBS!A:E, 5, FALSE)").number_format = "#,##0.00"
        
        # PV - SUMIFS on Fact_Baseline_PV up to Status Date
        ws_dash.cell(row=i, column=5, value=f"=SUMIFS(Fact_Baseline_PV!C:C, Fact_Baseline_PV!B:B, A{i}, Fact_Baseline_PV!A:A, \"<=\"&B$4)").number_format = "#,##0.00"
        
        # AC - SUMIFS on Fact_Actual_Costs
        ws_dash.cell(row=i, column=6, value=f"=SUMIFS(Fact_Actual_Costs!D:D, Fact_Actual_Costs!B:B, A{i})").number_format = "#,##0.00"
        
        # Progress % - MAXIFS on Fact_Physical_Progress up to Status Date (with compatibility prefix)
        ws_dash.cell(row=i, column=7, value=f"=IFERROR(_xlfn.MAXIFS(Fact_Physical_Progress!C:C, Fact_Physical_Progress!B:B, A{i}, Fact_Physical_Progress!A:A, \"<=\"&B$4), 0)").number_format = "0.0%"
        
        # EV - Budget * Progress %
        ws_dash.cell(row=i, column=8, value=f"=D{i}*G{i}").number_format = "#,##0.00"
        
        # CV - EV - AC
        ws_dash.cell(row=i, column=9, value=f"=H{i}-F{i}").number_format = "#,##0.00"
        
        # SV - EV - PV
        ws_dash.cell(row=i, column=10, value=f"=H{i}-E{i}").number_format = "#,##0.00"
        
        # CPI - EV / AC
        ws_dash.cell(row=i, column=11, value=f"=IF(F{i}>0, H{i}/F{i}, 1.00)").number_format = "0.00"
        
        # SPI - EV / PV
        ws_dash.cell(row=i, column=12, value=f"=IF(E{i}>0, H{i}/E{i}, 1.00)").number_format = "0.00"
        
        # EAC - BAC / CPI
        ws_dash.cell(row=i, column=13, value=f"=IF(K{i}>0, D{i}/K{i}, D{i})").number_format = "#,##0.00"
        
        # ETC - EAC - AC
        ws_dash.cell(row=i, column=14, value=f"=M{i}-F{i}").number_format = "#,##0.00"
        
        # VAC - BAC - EAC
        ws_dash.cell(row=i, column=15, value=f"=D{i}-M{i}").number_format = "#,##0.00"
        
        # Status - Color coded warnings if CPI < 0.95 or SPI < 0.90
        ws_dash.cell(row=i, column=16, value=f'=IF(K{i}<0.95, "🔴 Overrun", IF(K{i}<1.0, "🟡 Warning", "On Track"))').alignment = Alignment(horizontal="center")
        
        # Style row data cells
        for c in range(1, 17):
            cell = ws_dash.cell(row=i, column=c)
            cell.font = data_font
            cell.border = border_data
            if c not in [1, 2, 3, 16]:
                cell.alignment = Alignment(horizontal="right")
            
    # Add Total Row
    tot_row = 30
    ws_dash.cell(row=tot_row, column=1, value="Total / Project Portfolio").font = bold_data_font
    ws_dash.cell(row=tot_row, column=1).alignment = Alignment(horizontal="left")
    
    # Formulas for totals
    ws_dash.cell(row=tot_row, column=4, value="=SUM(D18:D29)").number_format = "#,##0.00" # BAC
    ws_dash.cell(row=tot_row, column=5, value="=SUM(E18:E29)").number_format = "#,##0.00" # PV
    ws_dash.cell(row=tot_row, column=6, value="=SUM(F18:F29)").number_format = "#,##0.00" # AC
    ws_dash.cell(row=tot_row, column=7, value="=H30/D30").number_format = "0.0%"       # Progress % (EV/BAC)
    ws_dash.cell(row=tot_row, column=8, value="=SUM(H18:H29)").number_format = "#,##0.00" # EV
    ws_dash.cell(row=tot_row, column=9, value="=H30-F30").number_format = "#,##0.00"     # CV
    ws_dash.cell(row=tot_row, column=10, value="=H30-E30").number_format = "#,##0.00"    # SV
    ws_dash.cell(row=tot_row, column=11, value="=IF(F30>0, H30/F30, 1.00)").number_format = "0.00" # CPI
    ws_dash.cell(row=tot_row, column=12, value="=IF(E30>0, H30/E30, 1.00)").number_format = "0.00" # SPI
    ws_dash.cell(row=tot_row, column=13, value="=IF(K30>0, D30/K30, D30)").number_format = "#,##0.00" # EAC
    ws_dash.cell(row=tot_row, column=14, value="=M30-F30").number_format = "#,##0.00"    # ETC
    ws_dash.cell(row=tot_row, column=15, value="=D30-M30").number_format = "#,##0.00"    # VAC
    
    # Status formula
    ws_dash.cell(row=tot_row, column=16, value=f'=IF(K{tot_row}<0.95, "🔴 Overrun", IF(K{tot_row}<1.0, "Warning", "On Track"))').alignment = Alignment(horizontal="center")
    
    for c in range(1, 17):
        cell = ws_dash.cell(row=tot_row, column=c)
        cell.font = bold_data_font
        cell.fill = total_fill
        cell.border = border_total
        if c not in [1, 2, 3, 16]:
            cell.alignment = Alignment(horizontal="right")
            
    # ----------------------------------------------------
    # Weekly Performance Tracking Table (Row 33+)
    # ----------------------------------------------------
    ws_dash["A33"] = "Weekly Cumulative Performance Trends"
    ws_dash["A33"].font = section_font
    
    trend_headers = [
        "Week", "Week Date", "Weekly PV", "Cum PV", "Weekly AC", "Cum AC", "Cum EV", "CV", "SV", "CPI", "SPI"
    ]
    
    for idx, h in enumerate(trend_headers, start=1):
        cell = ws_dash.cell(row=35, column=idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_header
        cell.alignment = Alignment(horizontal="left" if idx == 2 else "right")
        
    weeks_dates = [
        (1, "2026-05-04"), (2, "2026-05-11"), (3, "2026-05-18"), (4, "2026-05-25"),
        (5, "2026-06-01"), (6, "2026-06-08"), (7, "2026-06-15"), (8, "2026-06-22"),
        (9, "2026-06-29"), (10, "2026-07-06"), (11, "2026-07-13"), (12, "2026-07-20")
    ]
    
    for i, (wk, date_str) in enumerate(weeks_dates, start=36):
        ws_dash.cell(row=i, column=1, value=wk).alignment = Alignment(horizontal="right")
        
        date_cell = ws_dash.cell(row=i, column=2, value=datetime.strptime(date_str, "%Y-%m-%d").date())
        date_cell.alignment = Alignment(horizontal="left")
        date_cell.number_format = "yyyy-mm-dd"
        
        # Weekly PV - Sum PV for that week
        ws_dash.cell(row=i, column=3, value=f"=SUMIFS(Fact_Baseline_PV!C:C, Fact_Baseline_PV!A:A, B{i})").number_format = "#,##0.00"
        
        # Cum PV - Cumulative sum of Weekly PV (starts at row 36)
        ws_dash.cell(row=i, column=4, value=f"=SUM(C$36:C{i})").number_format = "#,##0.00"
        
        # Weekly AC - Sum AC for week dates
        ws_dash.cell(row=i, column=5, value=f"=SUMIFS(Fact_Actual_Costs!D:D, Fact_Actual_Costs!A:A, \">=\"&B{i}, Fact_Actual_Costs!A:A, \"<=\"&B{i}+6)").number_format = "#,##0.00"
        
        # Cum AC - Cumulative sum of Weekly AC (starts at row 36)
        ws_dash.cell(row=i, column=6, value=f"=SUM(E$36:E{i})").number_format = "#,##0.00"
        
        # Cum EV - Sum over WBS of Budget * Progress % at this week
        ws_dash.cell(row=i, column=7, value=f"=SUMPRODUCT(Dim_WBS!$E$2:$E$13, SUMIFS(Fact_Physical_Progress!$C$2:$C$100, Fact_Physical_Progress!$B$2:$B$100, Dim_WBS!$A$2:$A$13, Fact_Physical_Progress!$A$2:$A$100, \"<=\"&B{i}+4))").number_format = "#,##0.00"
        
        # CV - Cum EV - Cum AC
        ws_dash.cell(row=i, column=8, value=f"=G{i}-F{i}").number_format = "#,##0.00"
        
        # SV - Cum EV - Cum PV
        ws_dash.cell(row=i, column=9, value=f"=G{i}-D{i}").number_format = "#,##0.00"
        
        # CPI - Cum EV / Cum AC
        ws_dash.cell(row=i, column=10, value=f"=IF(F{i}>0, G{i}/F{i}, 1.00)").number_format = "0.00"
        
        # SPI - Cum EV / Cum PV
        ws_dash.cell(row=i, column=11, value=f"=IF(D{i}>0, G{i}/D{i}, 1.00)").number_format = "0.00"
        
        # Styles
        for c in range(1, 12):
            cell = ws_dash.cell(row=i, column=c)
            cell.font = data_font
            cell.border = border_data
            
    # Auto-fit columns on Dashboard
    ws_dash.column_dimensions["A"].width = 12
    ws_dash.column_dimensions["B"].width = 28 # Activity Name
    ws_dash.column_dimensions["C"].width = 15 # Type
    ws_dash.column_dimensions["D"].width = 16 # BAC
    ws_dash.column_dimensions["E"].width = 20 # PV
    ws_dash.column_dimensions["F"].width = 16 # AC
    ws_dash.column_dimensions["G"].width = 14 # Progress %
    ws_dash.column_dimensions["H"].width = 18 # EV
    ws_dash.column_dimensions["I"].width = 14 # CV
    ws_dash.column_dimensions["J"].width = 14 # SV
    ws_dash.column_dimensions["K"].width = 10 # CPI
    ws_dash.column_dimensions["L"].width = 10 # SPI
    ws_dash.column_dimensions["M"].width = 16 # EAC
    ws_dash.column_dimensions["N"].width = 16 # ETC
    ws_dash.column_dimensions["O"].width = 16 # VAC
    ws_dash.column_dimensions["P"].width = 14 # Status
    
    # Save output
    wb.save(OUTPUT_PATH)
    print(f"Workbook successfully saved to {OUTPUT_PATH}")

if __name__ == "__main__":
    build_workbook()
