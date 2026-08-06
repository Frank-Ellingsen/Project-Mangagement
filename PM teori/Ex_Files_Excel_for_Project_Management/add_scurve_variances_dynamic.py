import openpyxl
from openpyxl.chart.reference import Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import range_boundaries

def add_scurve_variances_dynamic(wb_path):
    print(f"\nProcessing workbook: {wb_path}")
    wb = openpyxl.load_workbook(wb_path)
    if "EVM Charts" not in wb.sheetnames:
        print("EVM Charts sheet not found.")
        return
        
    ws = wb["EVM Charts"]
    if "tbl_Charts_SCurveHelper" not in ws.tables:
        print("tbl_Charts_SCurveHelper table not found.")
        return
        
    t = ws.tables["tbl_Charts_SCurveHelper"]
    min_col, min_row, max_col, max_row = range_boundaries(t.ref)
    print(f"Current table range: {t.ref}")
    
    # We want to add CV and SV in the next two columns (columns max_col+1 and max_col+2)
    col_cv = max_col + 1
    col_sv = max_col + 2
    
    letter_cv = openpyxl.utils.get_column_letter(col_cv)
    letter_sv = openpyxl.utils.get_column_letter(col_sv)
    
    ws.cell(row=min_row, column=col_cv, value="CV")
    ws.cell(row=min_row, column=col_sv, value="SV")
    
    # Formulas for body
    for r in range(min_row + 1, max_row + 1):
        ws.cell(row=r, column=col_cv, value=f"=E{r}-D{r}")  # Cum EV - Cum AC
        ws.cell(row=r, column=col_sv, value=f"=E{r}-C{r}")  # Cum EV - Cum PV
        
    # Update table range ref
    new_ref = f"{openpyxl.utils.get_column_letter(min_col)}{min_row}:{letter_sv}{max_row}"
    t.ref = new_ref
    print(f"Updated table range: {new_ref}")
    
    # Add series to LineChart (Chart 1)
    if len(ws._charts) > 1:
        chart = ws._charts[1]
        
        # Add the two columns
        data_ref = Reference(ws, min_col=col_cv, min_row=min_row, max_col=col_sv, max_row=max_row)
        chart.add_data(data_ref, titles_from_data=True)
        print("Added CV and SV series to S-curve LineChart.")
        
        # Enable data labels on the last two series
        if len(chart.series) >= 2:
            s_cv = chart.series[-2]
            s_sv = chart.series[-1]
            
            s_cv.dLbls = DataLabelList()
            s_cv.dLbls.showVal = True
            
            s_sv.dLbls = DataLabelList()
            s_sv.dLbls.showVal = True
            print("Enabled data labels for CV and SV series.")
            
    wb.save(wb_path)
    print("Saved changes successfully.")

# Run it on populated test file
add_scurve_variances_dynamic(r"C:\Users\frank\Desktop\Project Mng\PM teori\Ex_Files_Excel_for_Project_Management\Test_Master_Mock_Populated_v3.xlsx")
