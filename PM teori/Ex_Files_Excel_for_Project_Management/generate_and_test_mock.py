import os
import csv
import openpyxl
from openpyxl.utils import range_boundaries

wb_source_path = r"C:\Users\frank\Desktop\Project Mng\PM teori\Ex_Files_Excel_for_Project_Management\Project_Controlling_App3.xlsx"
template_path = r"C:\Users\frank\Desktop\Project Mng\PM teori\Ex_Files_Excel_for_Project_Management\Master_Project_Controlling_Template.xlsx"
output_path = r"C:\Users\frank\Desktop\Project Mng\PM teori\Ex_Files_Excel_for_Project_Management\Test_Master_Mock_Populated_v2.xlsx"
mock_dir = r"C:\Users\frank\Desktop\Project Mng\PM teori\Ex_Files_Excel_for_Project_Management\mock_data_test"

os.makedirs(mock_dir, exist_ok=True)

# 1. Custom mock data for the project PRJ999: Shipyard Assembly Beta
custom_mock_data = {
    "tbl_Dim_Project": [
        ["Project_ID", "Project_Name", "Start_Date", "End_Date", "BAC", "Sector"],
        ["PRJ999", "Shipyard Assembly Beta", "2026-09-01", "2026-10-31", 60000.0, "Maritime"]
    ],
    "tbl_Dim_WBS": [
        ["WBS_ID", "WBS_Code", "Activity_Name", "Type", "Total_Budget"],
        ["WBS001", "1.0", "Design Phase", "Summary Task", 20000.0],
        ["WBS002", "1.1", "Structural Steel Work", "Activity", 25000.0],
        ["WBS003", "1.2", "Electrical & Outfitting", "Activity", 15000.0]
    ],
    "tbl_Dim_Resource": [
        ["Resource_ID", "Resource_Name", "Role", "Hourly_Rate"],
        ["RES101", "Alex Smith", "Structural Engineer", 200.0],
        ["RES102", "Sarah Jones", "Electrician", 150.0]
    ],
    "tbl_Fact_Baseline_PV": [
        ["Date", "WBS_ID", "Planned_Value"],
        ["2026-09-07", "WBS001", 5000.0],
        ["2026-09-14", "WBS001", 15000.0],
        ["2026-09-21", "WBS002", 10000.0],
        ["2026-09-28", "WBS002", 15000.0],
        ["2026-10-05", "WBS003", 5000.0],
        ["2026-10-12", "WBS003", 10000.0]
    ],
    "tbl_Fact_Actual_Costs": [
        ["Date", "WBS_ID", "Resource_ID", "Actual_Cost", "Hours_Worked"],
        ["2026-09-07", "WBS001", "RES101", 4500.0, 22.5],
        ["2026-09-14", "WBS001", "RES101", 16000.0, 80.0],
        ["2026-09-21", "WBS002", "RES101", 11000.0, 55.0],
        ["2026-09-28", "WBS002", "RES102", 12000.0, 80.0]
    ],
    "tbl_Fact_Physical_Progress": [
        ["Date", "WBS_ID", "Physical_Progress_Pct"],
        ["2026-09-07", "WBS001", 0.25],
        ["2026-09-14", "WBS001", 1.00],
        ["2026-09-21", "WBS002", 0.40],
        ["2026-09-28", "WBS002", 0.90]
    ],
    "tbl_Review_ActionLog": [
        ["Action ID", "Date Raised", "Description", "Owner", "Severity", "Status"],
        ["ACT-01", "2026-09-10", "Verify steel delivery delay impact", "Alex Smith", "High", "Open"]
    ],
    "tbl_PID_ProjectControls": [
        ["Role", "Description", "Name", "Contact", "Approval Status"],
        ["Project Controller", "Cost and Schedule control Lead", "Frank Ellingsen", "frank@company.com", "Approved"]
    ],
    "tbl_PID_RiskControl": [
        ["Risk ID", "Risk Description", "Probability", "Impact", "Mitigation Strategy"],
        ["RSK-01", "Steel Price Fluctuation", "Medium", "High", "Fixed-price supply contracts"]
    ],
    "tbl_PID_Financials": [
        ["Cost Category", "Baseline Budget", "Committed Costs", "Forecasted Margin", "Comments", "Status"],
        ["Labor", 35000.0, 33500.0, 1500.0, "On track", "Green"],
        ["Materials", 25000.0, 25000.0, 0.0, "Fully committed", "Green"]
    ],
    "tbl_Task_Plan": [
        ["WBS_ID", "Task Name", "Start Date", "End Date", "Duration", "Predecessors", "Resources", "Status", "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"],
        ["WBS001", "Design Phase", "2026-09-01", "2026-09-14", 14, "", "RES101", "Complete", None, None, None, None, None, None, None, None, None, None, None, None],
        ["WBS002", "Structural Steel Work", "2026-09-15", "2026-09-28", 14, "WBS001", "RES101", "In Progress", None, None, None, None, None, None, None, None, None, None, None, None],
        ["WBS003", "Electrical & Outfitting", "2026-09-29", "2026-10-12", 14, "WBS002", "RES102", "Not Started", None, None, None, None, None, None, None, None, None, None, None, None]
    ],
    "tbl_Task_Capture": [
        ["Task ID", "Date", "Resource", "Actual Hours", "Actual Cost", "Progress Pct", "Status", "Notes", "A", "B"],
        ["WBS001", "2026-09-07", "RES101", 22.5, 4500.0, 0.25, "In Progress", "Design initial drafting", None, None],
        ["WBS001", "2026-09-14", "RES101", 80.0, 16000.0, 1.00, "Complete", "Approved by client", None, None]
    ],
    "tbl_PERT_Data": [
        ["Activity ID", "Description", "Optimistic", "Most Likely", "Pessimistic", "Expected", "Predecessors", "ES", "EF", "Slack"],
        ["A1", "Conceptual Design", 5.0, 7.0, 9.0, 7.0, "", 0.0, 7.0, 0.0],
        ["A2", "Detail Drafting", 4.0, 7.0, 10.0, 7.0, "A1", 7.0, 14.0, 0.0]
    ],
    "tbl_Red_Action_Recommendations": [
        ["ID", "Issue Description", "EVM Indicator", "Impact Level", "Recommended Action", "Owner", "Target Date", "Status"],
        ["REC-01", "Design Phase Actual Cost slightly overran baseline", "CPI: 0.98", "Low", "Monitor labor rates", "Frank Ellingsen", "2026-09-15", "Closed"]
    ]
}

# 2. Extract the remaining tables from source workbook to generate mock CSVs
print("Reading source workbook to extract all remaining tables...")
wb_src = openpyxl.load_workbook(wb_source_path, data_only=True)

all_tables_data = {}

for sheet_name in wb_src.sheetnames:
    ws = wb_src[sheet_name]
    for t in ws.tables.values():
        if t.name in custom_mock_data:
            all_tables_data[t.name] = custom_mock_data[t.name]
        else:
            # Extract data from source table
            min_col, min_row, max_col, max_row = range_boundaries(t.ref)
            table_rows = []
            for r in ws[t.ref]:
                table_rows.append([cell.value for cell in r])
            all_tables_data[t.name] = table_rows

# Write all tables to CSV in mock directory
for name, rows in all_tables_data.items():
    csv_file = os.path.join(mock_dir, f"{name}.csv")
    with open(csv_file, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(rows)
    print(f"  Saved CSV: {name}.csv")

# 3. Load Master Template and populate ALL tables (from all tabs)
print("\nLoading Master Template...")
wb_template = openpyxl.load_workbook(template_path)

def coerce_value(value: str):
    if value is None or value == "":
        return None
    val_strip = str(value).strip()
    try:
        if "." in val_strip:
            return float(val_strip)
        return int(val_strip)
    except ValueError:
        return val_strip

populated_tables_count = 0

for sheet_name in wb_template.sheetnames:
    ws = wb_template[sheet_name]
    tables = list(ws.tables.values())
    if tables:
        print(f"Processing sheet: {sheet_name}")
        for t in tables:
            csv_file = os.path.join(mock_dir, f"{t.name}.csv")
            if os.path.exists(csv_file):
                with open(csv_file, "r", encoding="utf-8-sig") as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                
                csv_body = rows[1:]
                min_col, min_row, max_col, max_row = range_boundaries(t.ref)
                
                # Clear existing table data in the template
                for r in range(min_row + 1, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        ws.cell(row=r, column=c, value=None)
                
                # Write CSV data into table body
                for r_idx, row_data in enumerate(csv_body):
                    target_row = min_row + 1 + r_idx
                    for c_idx, val in enumerate(row_data):
                        col_idx = min_col + c_idx
                        if col_idx <= max_col:
                            ws.cell(row=target_row, column=col_idx, value=coerce_value(val))
                
                # Update the table range bounds
                new_max_row = min_row + len(csv_body)
                if new_max_row == min_row:
                    new_max_row = min_row + 1
                
                new_ref = f"{openpyxl.utils.get_column_letter(min_col)}{min_row}:{openpyxl.utils.get_column_letter(max_col)}{new_max_row}"
                t.ref = new_ref
                print(f"  Populated Table: {t.name} -> {new_ref} ({len(csv_body)} rows)")
                populated_tables_count += 1

print("\nSaving test workbook...")
wb_template.save(output_path)
print(f"Successfully populated all {populated_tables_count} tables in the workbook!")
print(f"Output saved to: {output_path}")
