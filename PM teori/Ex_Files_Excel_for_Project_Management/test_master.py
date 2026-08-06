import openpyxl
from openpyxl.utils import range_boundaries
import os
import csv

template_path = r"C:\Users\frank\Desktop\Project Mng\PM teori\Ex_Files_Excel_for_Project_Management\Master_Project_Controlling_Template.xlsx"
csv_dir = r"C:\Users\frank\Desktop\Project Mng\PM teori\Ex_Files_Excel_for_Project_Management\extracted_tables"
output_path = r"C:\Users\frank\Desktop\Project Mng\PM teori\Ex_Files_Excel_for_Project_Management\Test_Master_Populated.xlsx"

print("Loading template...")
wb = openpyxl.load_workbook(template_path)

# List of tables to populate from CSVs
tables_to_populate = [
    "tbl_Review_ActionLog",
    "tbl_PID_ProjectControls",
    "tbl_PID_RiskControl",
    "tbl_PID_Financials",
    "tbl_Task_Plan",
    "tbl_Task_Capture",
    "tbl_Red_Action_Recommendations",
    "tbl_Dim_Project",
    "tbl_Dim_WBS",
    "tbl_Dim_Resource",
    "tbl_Fact_Baseline_PV",
    "tbl_Fact_Actual_Costs",
    "tbl_Fact_Physical_Progress",
    "tbl_PERT_Data"
]

def coerce_value(value: str):
    if value is None or value == "":
        return None
    val_strip = str(value).strip()
    # Try converting numeric values
    try:
        if "." in val_strip:
            return float(val_strip)
        return int(val_strip)
    except ValueError:
        return val_strip

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    tables = list(ws.tables.values())
    for t in list(tables):
        if t.name in tables_to_populate:
            csv_file = os.path.join(csv_dir, f"{t.name}.csv")
            if not os.path.exists(csv_file):
                print(f"Warning: CSV file not found for {t.name}: {csv_file}")
                continue
                
            print(f"Populating table {t.name} from {csv_file}...")
            
            # Read CSV data (skipping the header)
            with open(csv_file, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                rows = list(reader)
            
            csv_header = rows[0]
            csv_body = rows[1:]
            
            # Get table boundaries
            min_col, min_row, max_col, max_row = range_boundaries(t.ref)
            
            # Clear all current body rows in the table first
            # The body in the template currently has 1 row (at min_row + 1)
            for r in range(min_row + 1, max_row + 1):
                for c in range(min_col, max_col + 1):
                    ws.cell(row=r, column=c, value=None)
            
            # Write CSV data into sheet
            for r_idx, row_data in enumerate(csv_body):
                target_row = min_row + 1 + r_idx
                for c_idx, val in enumerate(row_data):
                    col_idx = min_col + c_idx
                    if col_idx <= max_col:
                        ws.cell(row=target_row, column=col_idx, value=coerce_value(val))
            
            # Update the table reference (ref) to fit the new data size
            new_max_row = min_row + len(csv_body)
            # Table must have at least 1 header and 1 body row
            if new_max_row == min_row:
                new_max_row = min_row + 1
            
            new_ref = f"{openpyxl.utils.get_column_letter(min_col)}{min_row}:{openpyxl.utils.get_column_letter(max_col)}{new_max_row}"
            t.ref = new_ref
            print(f"  Table updated range: {new_ref} ({len(csv_body)} data rows)")

print("\nSaving populated test file...")
wb.save(output_path)
print(f"Test master file saved successfully to: {output_path}")
