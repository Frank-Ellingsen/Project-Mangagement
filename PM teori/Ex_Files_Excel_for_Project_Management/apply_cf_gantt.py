import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.formatting.rule import FormulaRule

template_path = r"C:\Users\frank\Desktop\Project Mng\PM teori\Ex_Files_Excel_for_Project_Management\Master_Project_Controlling_Template.xlsx"
source_populated_path = r"C:\Users\frank\Desktop\Project Mng\PM teori\Ex_Files_Excel_for_Project_Management\Test_Master_Mock_Populated_v2.xlsx"
populated_path = r"C:\Users\frank\Desktop\Project Mng\PM teori\Ex_Files_Excel_for_Project_Management\Test_Master_Mock_Populated_v3.xlsx"

def apply_timeline_formatting(wb_path, save_path=None):
    if save_path is None:
        save_path = wb_path
    print(f"Applying vertical timeline formatting to: {wb_path} -> saving to {save_path}")
    wb = openpyxl.load_workbook(wb_path)
    
    if "Simple Gantt Model" in wb.sheetnames:
        ws = wb["Simple Gantt Model"]
        
        # 1. Styles for Status Date Week
        red_dotted = Side(border_style="dotted", color="FF0000")
        border_status = Border(left=red_dotted, right=red_dotted)
        fill_status = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid") # light red tint
        
        # 2. Styles for Milestones (cell contains "M" or "◆")
        fill_milestone = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid") # light gold
        font_milestone = Font(color="F39C12", bold=True)
        border_milestone = Border(
            left=Side(border_style="thin", color="F39C12"),
            right=Side(border_style="thin", color="F39C12"),
            top=Side(border_style="thin", color="F39C12"),
            bottom=Side(border_style="thin", color="F39C12")
        )

        # Formula for Status Week:
        # For Hours grid (C13:N24) checking row 11 headers:
        # Formula checks if status date falls in this week
        formula_status_hours = 'AND(Control_Status_Date>=C$11, Control_Status_Date<C$11+7)'
        rule_status_hours = FormulaRule(formula=[formula_status_hours], border=border_status, fill=fill_status)
        
        # For Costs grid (C31:N42) checking row 29 headers:
        formula_status_costs = 'AND(Control_Status_Date>=C$29, Control_Status_Date<C$29+7)'
        rule_status_costs = FormulaRule(formula=[formula_status_costs], border=border_status, fill=fill_status)
        
        # Formula for Milestones (cell equals "M" or contains "◆"):
        rule_ms = FormulaRule(formula=['OR(C13="M", C13="◆")'], fill=fill_milestone, font=font_milestone, border=border_milestone)
        rule_ms_costs = FormulaRule(formula=['OR(C31="M", C31="◆")'], fill=fill_milestone, font=font_milestone, border=border_milestone)

        # Clear existing rules if we want to rebuild or just append
        # Let's add them at priority 1 (top of list) so they render correctly
        ws.conditional_formatting.add("C13:N24", rule_status_hours)
        ws.conditional_formatting.add("C31:N42", rule_status_costs)
        ws.conditional_formatting.add("C13:N24", rule_ms)
        ws.conditional_formatting.add("C31:N42", rule_ms_costs)
        
        print("  Added conditional formatting rules to 'Simple Gantt Model' sheet successfully.")
        
    wb.save(save_path)

apply_timeline_formatting(template_path)
apply_timeline_formatting(source_populated_path, populated_path)
print("Finished applying all Gantt enhancements!")
