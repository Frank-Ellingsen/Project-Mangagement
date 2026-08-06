import os
import csv
import openpyxl
from openpyxl.utils import range_boundaries
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.chart.reference import Reference
from openpyxl.chart.label import DataLabelList

# Define paths
ROOT_DIR = r"C:\Users\frank\Desktop\Project Mng\PM teori\Ex_Files_Excel_for_Project_Management"
SRC_APP = os.path.join(ROOT_DIR, "Project_Controlling_App3.xlsx")
TEMPLATE_PATH = os.path.join(ROOT_DIR, "Master_Project_Controlling_Template.xlsx")
MOCK_DIR = os.path.join(ROOT_DIR, "mock_data_test")
OUTPUT_PATH = os.path.join(ROOT_DIR, "Test_Master_Mock_Populated_Final.xlsx")

os.makedirs(MOCK_DIR, exist_ok=True)

# 1. Setup Mock Datasets
custom_mock_data = {
    "tbl_Dim_Project": [
        ["Project_ID", "Project_Name", "Start_Date", "End_Date", "BAC", "Sector"],
        ["PRJ999", "Shipyard Assembly Beta", "2026-09-01", "2026-10-31", 60000.0, "Maritime"]
    ],
    "tbl_Dim_WBS": [
        ["WBS_ID", "WBS_Code", "Activity_Name", "Type", "Total_Budget"],
        ["WBS001", "1.0", "Design Phase", "Summary Task", 20000.0],
        ["WBS002", "1.1", "Structural Steel Work", "Activity", 25000.0],
        ["WBS003", "1.2", "Electrical & Outfitting", "Activity", 15000.0]
    ],
    "tbl_Dim_Resource": [
        ["Resource_ID", "Resource_Name", "Role", "Hourly_Rate"],
        ["RES101", "Alex Smith", "Structural Engineer", 200.0],
        ["RES102", "Sarah Jones", "Electrician", 150.0]
    ],
    "tbl_Fact_Baseline_PV": [
        ["Date", "WBS_ID", "Planned_Value"],
        ["2026-09-07", "WBS001", 5000.0],
        ["2026-09-14", "WBS001", 15000.0],
        ["2026-09-21", "WBS002", 10000.0],
        ["2026-09-28", "WBS002", 15000.0],
        ["2026-10-05", "WBS003", 5000.0],
        ["2026-10-12", "WBS003", 10000.0]
    ],
    "tbl_Fact_Actual_Costs": [
        ["Date", "WBS_ID", "Resource_ID", "Actual_Cost", "Hours_Worked"],
        ["2026-09-07", "WBS001", "RES101", 4500.0, 22.5],
        ["2026-09-14", "WBS001", "RES101", 16000.0, 80.0],
        ["2026-09-21", "WBS002", "RES101", 11000.0, 55.0],
        ["2026-09-28", "WBS002", "RES102", 12000.0, 80.0]
    ],
    "tbl_Fact_Physical_Progress": [
        ["Date", "WBS_ID", "Physical_Progress_Pct"],
        ["2026-09-07", "WBS001", 0.25],
        ["2026-09-14", "WBS001", 1.00],
        ["2026-09-21", "WBS002", 0.40],
        ["2026-09-28", "WBS002", 0.90]
    ],
    "tbl_Review_ActionLog": [
        ["Action ID", "Date Raised", "Description", "Owner", "Severity", "Status"],
        ["ACT-01", "2026-09-10", "Verify steel delivery delay impact", "Alex Smith", "High", "Open"]
    ],
    "tbl_PID_ProjectControls": [
        ["Role", "Description", "Name", "Contact", "Approval Status"],
        ["Project Controller", "Cost and Schedule control Lead", "Frank Ellingsen", "frank@company.com", "Approved"]
    ],
    "tbl_PID_RiskControl": [
        ["Risk ID", "Risk Description", "Probability", "Impact", "Mitigation Strategy"],
        ["RSK-01", "Steel Price Fluctuation", "Medium", "High", "Fixed-price supply contracts"]
    ],
    "tbl_PID_Financials": [
        ["Cost Category", "Baseline Budget", "Committed Costs", "Forecasted Margin", "Comments", "Status"],
        ["Labor", 35000.0, 33500.0, 1500.0, "On track", "Green"],
        ["Materials", 25000.0, 25000.0, 0.0, "Fully committed", "Green"]
    ],
    "tbl_Task_Plan": [
        ["WBS_ID", "Task Name", "Start Date", "End Date", "Duration", "Predecessors", "Resources", "Status", "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"],
        ["WBS001", "Design Phase", "2026-09-01", "2026-09-14", 14, "", "RES101", "Complete", None, None, None, None, None, None, None, None, None, None, None, None],
        ["WBS002", "Structural Steel Work", "2026-09-15", "2026-09-28", 14, "WBS001", "RES101", "In Progress", None, None, None, None, None, None, None, None, None, None, None, None],
        ["WBS003", "Electrical & Outfitting", "2026-09-29", "2026-10-12", 14, "WBS002", "RES102", "Not Started", None, None, None, None, None, None, None, None, None, None, None, None]
    ],
    "tbl_Task_Capture": [
        ["Task ID", "Date", "Resource", "Actual Hours", "Actual Cost", "Progress Pct", "Status", "Notes", "A", "B"],
        ["WBS001", "2026-09-07", "RES101", 22.5, 4500.0, 0.25, "In Progress", "Design initial drafting", None, None],
        ["WBS001", "2026-09-14", "RES101", 80.0, 16000.0, 1.00, "Complete", "Approved by client", None, None]
    ],
    "tbl_PERT_Data": [
        ["Activity ID", "Description", "Optimistic", "Most Likely", "Pessimistic", "Expected", "Predecessors", "ES", "EF", "Slack"],
        ["A1", "Conceptual Design", 5.0, 7.0, 9.0, 7.0, "", 0.0, 7.0, 0.0],
        ["A2", "Detail Drafting", 4.0, 7.0, 10.0, 7.0, "A1", 7.0, 14.0, 0.0]
    ],
    "tbl_Red_Action_Recommendations": [
        ["ID", "Issue Description", "EVM Indicator", "Impact Level", "Recommended Action", "Owner", "Target Date", "Status"],
        ["REC-01", "Design Phase Actual Cost slightly overran baseline", "CPI: 0.98", "Low", "Monitor labor rates", "Frank Ellingsen", "2026-09-15", "Closed"]
    ]
}

def run_pipeline():
    print(">>> Starting Project Controlling App Automation Pipeline...")
    
    # 1. Extract non-customized tables from App3.xlsx to mock directory
    print("Step 1: Extracting structural sheets...")
    wb_src = openpyxl.load_workbook(SRC_APP, data_only=True)
    all_data = {}
    for sheet_name in wb_src.sheetnames:
        ws = wb_src[sheet_name]
        for t in ws.tables.values():
            if t.name in custom_mock_data:
                all_data[t.name] = custom_mock_data[t.name]
            else:
                min_col, min_row, max_col, max_row = range_boundaries(t.ref)
                table_rows = []
                for r in ws[t.ref]:
                    table_rows.append([cell.value for cell in r])
                all_data[t.name] = table_rows
                
    for name, rows in all_data.items():
        csv_file = os.path.join(MOCK_DIR, f"{name}.csv")
        with open(csv_file, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerows(rows)

    # 2. Generalize Master Excel Template
    print("Step 2: Resetting Master Template to a clean, generic state...")
    wb_temp = openpyxl.load_workbook(SRC_APP)
    tables_to_clean = {
        "tbl_Review_ActionLog": [1, "2026-01-01", "Action description", "Owner", "Medium", "Open"],
        "tbl_PID_ProjectControls": ["Project Controls Lead", "Role Description", "Name", "Contact", "Approved"],
        "tbl_PID_RiskControl": ["Risk ID", "Risk Description", "Probability", "Impact", "Mitigation Strategy"],
        "tbl_PID_Financials": ["Cost Category", 0.0, 0.0, 0.0, "Comments", "Status"],
        "tbl_Task_Plan": ["WBS_ID", "Task Name", "Start Date", "End Date", "Duration", "Predecessors", "Resources", "Status"] + [None]*12,
        "tbl_Task_Capture": ["Task ID", "Date", "Resource", "Actual Hours", "Actual Cost", "Progress Pct", "Status", "Notes", None, None],
        "tbl_Dim_Project": ["PRJ001", "Generic Master Project", "2026-01-01", "2026-12-31", 100000.0, "Defense/Engineering"],
        "tbl_Dim_WBS": ["WBS001", "1.0", "Project Management", "Summary Task", 10000.0],
        "tbl_Dim_Resource": ["RES001", "Resource Name", "Role", 150.0],
        "tbl_Fact_Baseline_PV": ["2026-01-07", "WBS001", 1000.0],
        "tbl_Fact_Actual_Costs": ["2026-01-07", "WBS001", "RES001", 1200.0, 8.0],
        "tbl_Fact_Physical_Progress": ["2026-01-07", "WBS001", 0.10],
        "tbl_PERT_Data": ["Activity ID", "Description", 1.0, 2.0, 3.0, 2.0, "Predecessors", "ES", "EF", "Slack"],
        "tbl_Red_Action_Recommendations": ["ID", "Issue Description", "EVM Indicator", "Impact Level", "Recommended Action", "Owner", "Target Date", "Status"]
    }
    
    for sheet_name in wb_temp.sheetnames:
        ws = wb_temp[sheet_name]
        for t in list(ws.tables.values()):
            if t.name in tables_to_clean:
                placeholder = tables_to_clean[t.name]
                min_col, min_row, max_col, max_row = range_boundaries(t.ref)
                
                # Overwrite first row
                for col_idx, val in enumerate(placeholder[:max_col - min_col + 1], start=min_col):
                    ws.cell(row=min_row + 1, column=col_idx, value=val)
                    
                # Clear other rows
                if max_row > min_row + 1:
                    for r in range(min_row + 2, max_row + 1):
                        for c in range(min_col, max_col + 1):
                            ws.cell(row=r, column=c, value=None)
                
                t.ref = f"{openpyxl.utils.get_column_letter(min_col)}{min_row}:{openpyxl.utils.get_column_letter(max_col)}{min_row + 1}"
    
    wb_temp.save(TEMPLATE_PATH)

    # 3. Build & Populate Populated Test Workbook
    print("Step 3: Populating Test Workbook from CSV datasets...")
    wb_pop = openpyxl.load_workbook(TEMPLATE_PATH)
    
    def coerce_val(value: str):
        if value is None or value == "":
            return None
        val_strip = str(value).strip()
        try:
            if "." in val_strip:
                return float(val_strip)
            return int(val_strip)
        except ValueError:
            return val_strip

    for sheet_name in wb_pop.sheetnames:
        ws = wb_pop[sheet_name]
        for t in list(ws.tables.values()):
            csv_file = os.path.join(MOCK_DIR, f"{t.name}.csv")
            if os.path.exists(csv_file):
                with open(csv_file, "r", encoding="utf-8-sig") as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                csv_body = rows[1:]
                min_col, min_row, max_col, max_row = range_boundaries(t.ref)
                
                # Clear template default row
                ws.cell(row=min_row+1, column=min_col, value=None)
                
                # Write CSV data
                for r_idx, row_data in enumerate(csv_body):
                    target_row = min_row + 1 + r_idx
                    for c_idx, val in enumerate(row_data):
                        col_idx = min_col + c_idx
                        if col_idx <= max_col:
                            ws.cell(row=target_row, column=col_idx, value=coerce_val(val))
                            
                new_max_row = min_row + len(csv_body)
                if new_max_row == min_row:
                    new_max_row = min_row + 1
                t.ref = f"{openpyxl.utils.get_column_letter(min_col)}{min_row}:{openpyxl.utils.get_column_letter(max_col)}{new_max_row}"

    # 4. Enhance Gantt Charts (Status Date Line & Milestones)
    print("Step 4: Adding Gantt Chart Enhancements (Conditional formatting)...")
    for wb_obj, path in [(wb_pop, OUTPUT_PATH), (openpyxl.load_workbook(TEMPLATE_PATH), TEMPLATE_PATH)]:
        ws_gantt = wb_obj["Simple Gantt Model"]
        
        red_dotted = Side(border_style="dotted", color="FF0000")
        border_status = Border(left=red_dotted, right=red_dotted)
        fill_status = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
        
        fill_milestone = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
        font_milestone = Font(color="F39C12", bold=True)
        border_milestone = Border(
            left=Side(border_style="thin", color="F39C12"),
            right=Side(border_style="thin", color="F39C12"),
            top=Side(border_style="thin", color="F39C12"),
            bottom=Side(border_style="thin", color="F39C12")
        )

        rule_status_hours = FormulaRule(formula=['AND(Control_Status_Date>=C$11, Control_Status_Date<C$11+7)'], border=border_status, fill=fill_status)
        rule_status_costs = FormulaRule(formula=['AND(Control_Status_Date>=C$29, Control_Status_Date<C$29+7)'], border=border_status, fill=fill_status)
        rule_ms = FormulaRule(formula=['OR(C13="M", C13="◆")'], fill=fill_milestone, font=font_milestone, border=border_milestone)
        rule_ms_costs = FormulaRule(formula=['OR(C31="M", C31="◆")'], fill=fill_milestone, font=font_milestone, border=border_milestone)

        ws_gantt.conditional_formatting.add("C13:N24", rule_status_hours)
        ws_gantt.conditional_formatting.add("C31:N42", rule_status_costs)
        ws_gantt.conditional_formatting.add("C13:N24", rule_ms)
        ws_gantt.conditional_formatting.add("C31:N42", rule_ms_costs)
        
        # 5. Enhance S-Curve (Add CV/SV helper columns and labels)
        print(f"Step 5: Enhancing EVM S-Curve Chart...")
        ws_charts = wb_obj["EVM Charts"]
        t_scurve = ws_charts.tables["tbl_Charts_SCurveHelper"]
        min_col, min_row, max_col, max_row = range_boundaries(t_scurve.ref)
        
        col_cv = max_col + 1
        col_sv = max_col + 2
        letter_cv = openpyxl.utils.get_column_letter(col_cv)
        letter_sv = openpyxl.utils.get_column_letter(col_sv)
        
        ws_charts.cell(row=min_row, column=col_cv, value="CV")
        ws_charts.cell(row=min_row, column=col_sv, value="SV")
        
        for r in range(min_row + 1, max_row + 1):
            ws_charts.cell(row=r, column=col_cv, value=f"=E{r}-D{r}")
            ws_charts.cell(row=r, column=col_sv, value=f"=E{r}-C{r}")
            
        t_scurve.ref = f"{openpyxl.utils.get_column_letter(min_col)}{min_row}:{letter_sv}{max_row}"
        
        if len(ws_charts._charts) > 1:
            chart = ws_charts._charts[1]
            data_ref = Reference(ws_charts, min_col=col_cv, min_row=min_row, max_col=col_sv, max_row=max_row)
            chart.add_data(data_ref, titles_from_data=True)
            
            if len(chart.series) >= 2:
                s_cv = chart.series[-2]
                s_sv = chart.series[-1]
                s_cv.dLbls = DataLabelList()
                s_cv.dLbls.showVal = True
                s_sv.dLbls = DataLabelList()
                s_sv.dLbls.showVal = True

        # 6. Compliance Enforcement (Controlling Best Practices)
        dashboard_sheets = ["EVM Dashboard", "EVM Charts", "Simple Gantt Model", "PERT Diagram"]
        for s_name in wb_obj.sheetnames:
            ws_item = wb_obj[s_name]
            if s_name in dashboard_sheets:
                ws_item.views.sheetView[0].showGridLines = False
            else:
                ws_item.views.sheetView[0].showGridLines = True

        if "EVM Dashboard" in wb_obj.sheetnames:
            ws_dash = wb_obj["EVM Dashboard"]
            for row in ws_dash.iter_rows(min_row=17, max_row=30, min_col=1, max_col=17):
                for cell in row:
                    cb = cell.border
                    cell.border = Border(top=cb.top if cb else None, bottom=cb.bottom if cb else None, left=None, right=None)
                    if cell.column in [1, 2, 3]:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif cell.column == 16:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="center")

            # Apply RAG conditional formatting on Column P (Status)
            green_f = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
            green_txt = Font(color="117A65", bold=True)
            rule_g = CellIsRule(operator='equal', formula=['"On Track"'], fill=green_f, font=green_txt)
            
            amber_f = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
            amber_txt = Font(color="D35400", bold=True)
            rule_a = CellIsRule(operator='equal', formula=['"Warning"'], fill=amber_f, font=amber_txt)
            
            red_f = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
            red_txt = Font(color="C0392B", bold=True)
            rule_r = CellIsRule(operator='equal', formula=['"Overrun"'], fill=red_f, font=red_txt)
            
            ws_dash.conditional_formatting.add("P18:P30", rule_g)
            ws_dash.conditional_formatting.add("P18:P30", rule_a)
            ws_dash.conditional_formatting.add("P18:P30", rule_r)

        wb_obj.save(path)
        
    print(f"\n>>> Automation Pipeline completed successfully! Final Populated workbook saved to: {OUTPUT_PATH}")

if __name__ == "__main__":
    run_pipeline()
